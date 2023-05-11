import csv
import io
import sys

import openpyxl as openpyxl
from flask import Response, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import and_
from sqlalchemy.exc import IntegrityError
from werkzeug.exceptions import BadRequest

from database import db
from database.models.rol import Rol
from database.models.test import AlumnosTest, Test
from database.models.user import User
from database.models.pregunta import Pregunta
from database.models.alumno import Alumno, AlumnosClase
from database.models.clase import Clase
from database.models.year import Year

from flask_restful import Resource

from resources.errors import SchemaValidationError, InternalServerError, AlumnoNotExistsError, AlumnoAlreadyExistsError
from datetime import datetime
from resources.decorators import profesor, alumno, superuser


class LoggedUserAlumno(Resource):
    @alumno
    def get(self):
        try:
            db.openSession()
            user_id = get_jwt_identity()
            user = db.session.query(Alumno, User, AlumnosTest.c.code, AlumnosClase.c.clase).join(AlumnosTest,
                    and_(AlumnosTest.c.id == user_id,Alumno.id == AlumnosTest.c.alumno, Alumno.activo == 1,
                    AlumnosTest.c.activo == 1)).join(User, Alumno.user == User.id).join(Test, Test.id == AlumnosTest.c.test).join(AlumnosClase,
                    AlumnosClase.c.alumno == Alumno.id).join(Clase, and_(Clase.id == AlumnosClase.c.clase, Test.clase == Clase.id)).one()
            user_json = user[1].serialize
            del user_json['id']
            del user_json['validado']
            alumno_json = user[0].serialize(user[3], user[2])
            alumno_json['alumno'] = alumno_json['id']
            del alumno_json['id']
            db.session.close()
            return jsonify({**user_json, **alumno_json})
        except AttributeError:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise AlumnoNotExistsError
        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise InternalServerError

class AlumnosApi(Resource):
    @profesor
    def get(self):
        try:
            db.openSession()
            user_id = get_jwt_identity()
            user = User.query.get(user_id)
            alumnos = db.session.query(Alumno, User).join(User, and_(User.id == Alumno.user,
                        User.colegio == user.colegio)).all()
            db.session.close()
            return jsonify(alumnos=[{**(i[0]).serialize(None), **i[1].serialize, 'id': i[0].id, 'clase_nombre': None} for i in alumnos])
        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise InternalServerError

    @profesor
    def post(self):
        try:
            db.openSession()
            body = request.get_json()
            user_id = get_jwt_identity()
            user_log = User.query.get(user_id)
            rol = Rol.query.filter_by(nombre="alumno").one().id
            date = datetime.now()
            if 'alumnos' in body:
                for al in body['alumnos']:
                    if 'colegio' not in al:
                        al['colegio'] = user_log.colegio
                    elif not al['colegio']:
                        al['colegio'] = user_log.colegio
                    if not al['alias']:
                        al['alias'] = al['nombre'] + " " + al['apellidos'][0:3]
                    if al['DNI']:
                        user = User(nombre=al['nombre'], apellidos=al['apellidos'], DNI=al['DNI'],
                                    colegio=al['colegio'], rol=rol, date_joined=date, fecha_nacimiento=al['fecha_nacimiento'])
                    else:
                        user = User(nombre=al['nombre'], apellidos=al['apellidos'],
                                    colegio=al['colegio'], rol=rol, date_joined=date, fecha_nacimiento=al['fecha_nacimiento'])
                    db.session.add(user)
                    db.session.commit()
                    if al['clase']:
                        clase = al['clase']
                    else:
                        clase = None
                    alumno = Alumno(sexo=al['sexo'],
                                    alias=al['alias'],
                                    user=user.id)
                    db.session.add(alumno)
                    db.session.commit()
                    if clase:
                        alumnoClase = AlumnosClase(alumno=alumno.id, clase=clase, alias=al['alias'])
                        db.session.add(alumnoClase)
                        db.session.commit()

            else:
                if 'colegio' not in body:
                    body['colegio'] = user_log.colegio
                elif not body['colegio']:
                    body['colegio'] = user_log.colegio
                if not body['alias']:
                    body['alias'] = body['nombre'] + " " + body['apellidos'][0:3]
                if body['DNI']:
                    user = User(nombre=body['nombre'], apellidos=body['apellidos'], DNI=body['DNI'],
                                colegio=body['colegio'], rol=rol, date_joined=date, fecha_nacimiento=body['fecha_nacimiento'])
                else:
                    user = User(nombre=body['nombre'], apellidos=body['apellidos'],
                                colegio=body['colegio'], rol=rol, date_joined=date, fecha_nacimiento=body['fecha_nacimiento'])
                db.session.add(user)
                db.session.commit()
                if body['clase']:
                    clase = body['clase']
                else:
                    clase = None
                alumno = Alumno(sexo=body['sexo'], alias=body['alias'],
                                user=user.id)
                db.session.add(alumno)
                db.session.commit()
                if clase:
                    alumnoClase = AlumnosClase(alumno=alumno.id, clase=clase, alias=body['alias'])
                    db.session.add(alumnoClase)
                    db.session.commit()
            db.session.close()
            return '', 200
        except IntegrityError:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise AlumnoAlreadyExistsError
        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise InternalServerError


class AlumnoApi(Resource):
    @profesor
    def put(self, id):
        try:
            db.openSession()
            alumno = db.session.query(Alumno).filter(Alumno.id == id).one()
            user = db.session.query(User).filter(User.id == alumno.user).one()
            body = request.get_json()
            year = Year.query.filter(Year.current == 1, Year.colegio == user.colegio).one()
            alumnoClase = db.session.query(AlumnosClase).join(Clase, and_(Clase.id == AlumnosClase.c.clase, year.id == Clase.year)).filter(AlumnosClase.c.alumno == id).first()
            if alumnoClase is not None:
                if 'alias' in body:
                    stmt = AlumnosClase.update().where(AlumnosClase.c.id == alumnoClase.id).values(alias=body['alias'])
                    db.session.execute(stmt)
            for key, value in body.items():
                if key in ("clase", "sexo", "alias", "activo"):
                    setattr(alumno, key, value)
                if key in ("username", "nombre", "apellidos", "DNI", "colegio", "password", "activo", "fecha_nacimiento"):
                    setattr(user, key, value)
                    if key == "password":
                        user.hash_password()
            if 'activo' in body:
                alumnos = db.session.query(AlumnosClase).filter(AlumnosClase.c.alumno == id).all()
                for al in alumnos:
                    update_stmnt = AlumnosTest.update().where(AlumnosTest.c.alumno==al.id).values(activo=body['activo'])
                    db.session.execute(update_stmnt)
            db.session.commit()
            db.session.close()
            return '', 200
        except AttributeError:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise AlumnoNotExistsError
        except IntegrityError:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise AlumnoAlreadyExistsError
        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise InternalServerError

    @profesor
    def delete(self, id):
        try:
            db.openSession()
            alumno = db.session.query(Alumno).filter_by(id = id).one()
            db.session.query(User).filter(User.id == alumno.user).delete()
            db.session.commit()
            db.session.close()
            return '', 200
        except AttributeError:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise AlumnoNotExistsError
        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise InternalServerError

    @profesor
    def get(self, id):
        try:
            db.openSession()
            alumno = db.session.query(Alumno, User).join(User,and_(User.id == Alumno.user, Alumno.id == id)).one()
            db.session.close()
            return jsonify({**(alumno[0]).serialize(None), **alumno[1].serialize})
        except AttributeError:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise AlumnoNotExistsError
        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise InternalServerError

class AlumnosApiSearch(Resource):
    @profesor
    def post(self):
        try:
            db.openSession()
            body = request.get_json()
            user_id = get_jwt_identity()
            user = User.query.get(user_id)
            if 'activo' not in body:
                body['activo'] = 1
            alumnos = db.session.query(Alumno, User, Clase.nombre, Clase.id).join(User, and_(User.id == Alumno.user, User.colegio == user.colegio)).join(AlumnosClase, AlumnosClase.c.alumno == Alumno.id, isouter=True).join(Clase, and_(Clase.id == AlumnosClase.c.clase, Clase.year == body['year']), isouter=True).filter(Alumno.activo == body['activo'])
            for attr, value in body.items():
                if attr.startswith('Alumno.'):
                    if value:
                        alumnos = alumnos.filter(getattr(Alumno, attr.replace('Alumno.', "")).like("%%%s%%" % value))
                    else:
                        alumnos = alumnos.filter(getattr(Alumno, attr.replace('Alumno.', "")).is_(value))
                else:
                    if attr != 'year' and attr != 'clase' and attr != 'activo':
                        alumnos = alumnos.filter(getattr(User, attr.replace('User.', "")).like("%%%s%%" % value))
            alumnos = alumnos.all()
            db.session.close()
            jsonObj = []
            for i in alumnos:
                if 'clase' in body:
                    if i[2] is None:
                        jsonObj.append({**(i[0]).serialize(i[3]), **i[1].serialize, 'id': i[0].id, 'clase_nombre': i[2]})
                else:
                    jsonObj.append({**(i[0]).serialize(i[3]), **i[1].serialize, 'id': i[0].id, 'clase_nombre': i[2]})
            return jsonify(alumnos=jsonObj)
        except AttributeError:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise AlumnoNotExistsError
        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            raise InternalServerError

class ImportExcel(Resource):
    @profesor
    def post(self):
        db.openSession()
        csvFileName = request.files['file']
        claseId = int(csvFileName.filename.split("-")[1].split(".")[0])
        if claseId == 0:
            clase = None
        else:
            clase = claseId

        if not csvFileName:
            raise BadRequest

        # decoded_file = csvFileName.read().decode('utf-8', errors='ignore').splitlines()
        # reader = csv.DictReader(decoded_file)
        errors = []
        try:
            user_id = get_jwt_identity()
            user_log = User.query.get(user_id)
            rol = Rol.query.filter_by(nombre="alumno").one().id
            dateNow = datetime.now()
            # print(reader)
            wb = openpyxl.load_workbook(csvFileName, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = {}
            for num, row in enumerate(ws.iter_rows()):
                if num == 0:
                    for n,cell in enumerate(row):
                        rows[cell.value] = n
                    if len(rows.items()) != 6 or 'DNI' not in rows or 'NOMBRE' not in rows or 'APELLIDO1' not in rows or 'APELLIDO2' not in rows or 'SEXO' not in rows or 'FECHA-NACIMIENTO' not in rows:
                        raise
                else:

                    if row[rows['DNI']].value:
                        if row[rows['DNI']].value != "":
                            dni = row[rows['DNI']].value
                        else:
                            dni = None
                    else:
                        dni = None
                    if isinstance(row[rows['FECHA-NACIMIENTO']].value, datetime):
                        fecha_nacimiento = row[rows['FECHA-NACIMIENTO']].value.strftime("%d-%m-%Y")
                    else:
                        if row[rows['FECHA-NACIMIENTO']].value:
                            date = row[rows['FECHA-NACIMIENTO']].value.split('/')
                            fecha_nacimiento = date[2] + '-' + date[1] + '-' + date[0]
                        else:
                            continue
                    if (row[rows['SEXO']].value == 'M'):
                        sexo = 'H'
                    elif (row[rows['SEXO']].value == 'F'):
                        sexo = 'M'
                    else:
                        sexo = 'M'
                    year = Year.query.filter(Year.current == 1, Year.colegio == user_log.colegio).one()
                    alumnoUserOld = None
                    newAlumnoUser = None
                    errorType = None
                    try:
                        if dni is not None:
                            alumnoUserOld = User.query.filter(User.DNI == dni).first()
                            errorType = 'dni'
                        if alumnoUserOld is None:
                            alumnoUserOld = User.query.filter(User.nombre == row[rows['NOMBRE']].value, User.apellidos==row[rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value, User.fecha_nacimiento==fecha_nacimiento).first()
                            errorType = 'date'
                        if alumnoUserOld is None:
                            newAlumnoUser = User(nombre=row[rows['NOMBRE']].value, apellidos=row[rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value, DNI=dni, rol=rol,
                                             colegio=user_log.colegio, fecha_nacimiento=fecha_nacimiento,  date_joined=dateNow)
                        else:
                            currentAlumno = db.session.query(Alumno).filter(Alumno.user==alumnoUserOld.id).one()
                            claseAlumno = db.session.query(AlumnosClase, Clase).join(Clase, and_(Clase.id == AlumnosClase.c.clase, AlumnosClase.c.alumno == currentAlumno.id, Clase.year == year.id)).first()
                            if claseAlumno:
                                raise
                            else:
                                insert_stmnt = AlumnosClase.insert().values(clase=clase, alumno=currentAlumno.id,
                                                                            alias=currentAlumno.alias)
                                try:
                                    db.session.execute(insert_stmnt)
                                    db.session.commit()
                                except:
                                    db.session.rollback()
                                    raise


                        try:
                            if alumnoUserOld is None:
                                db.session.add(newAlumnoUser)
                                db.session.commit()
                        except:
                            db.session.rollback()
                            raise

                        if alumnoUserOld is None:
                            newAlumno = Alumno(sexo=sexo, user=newAlumnoUser.id,
                                               alias=newAlumnoUser.nombre + " " + newAlumnoUser.apellidos[0:3])
                        try:
                            if alumnoUserOld is None:
                                db.session.add(newAlumno)
                                db.session.commit()
                        except:
                            db.session.rollback()
                            raise

                        try:
                            if alumnoUserOld is None:
                                insert_stmnt = AlumnosClase.insert().values(clase=clase, alumno=newAlumno.id,
                                                                        alias=newAlumno.alias)
                                db.session.execute(insert_stmnt)
                                db.session.commit()
                        except:
                            db.session.rollback()
                            raise
                    except RuntimeError as e:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
                        print("No se puede importar el alumno: " + row[rows['NOMBRE']].value + " " + row[
                            rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value + str(e), flush=True)
                        if errorType == 'dni':
                            errors.append("No se puede importar el alumno: " + row[rows['NOMBRE']].value + " " + row[
                                rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value + " DNI YA EXISTE (" +
                                          row[rows['DNI']].value + ")" + '\n')
                        elif errorType == 'date':
                            errors.append("No se puede añadir el alumno: " + row[rows['NOMBRE']].value + " " + row[
                                rows['APELLIDO1']].value + " " + row[
                                              rows['APELLIDO2']].value + " ALUMNO YA PERTENECE A OTRA CLASE" + '\n')
                    except Exception as e:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
                        print("No se puede importar el alumno: " + row[rows['NOMBRE']].value + " " + row[rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value + str(e), flush=True)
                        if 'DNI_UNIQUE' in str(e):
                            errors.append("No se puede importar el alumno: " + row[rows['NOMBRE']].value + " " + row[rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value + " DNI YA EXISTE (" + row[rows['DNI']].value + ")" + '\n')
                        elif 'alias_UNIQUE' in str(e):
                            errors.append("No se puede añadir el alumno: " + row[rows['NOMBRE']].value + " " + row[rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value + " ALIAS YA EXISTE EN ESTA CLASE (" + currentAlumno.alias if currentAlumno is not None else newAlumno.alias + ")" + '\n')
                        elif 'alumno_UNIQUE' in str(e):
                            errors.append("No se puede añadir el alumno: " + row[rows['NOMBRE']].value + " " + row[rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value + " ALUMNO YA EXISTE EN ESTA CLASE" + '\n')
                        else:
                            errors.append("No se puede importar el alumno: " + row[rows['NOMBRE']].value + " " + row[rows['APELLIDO1']].value + " " + row[rows['APELLIDO2']].value + " FECHA NACIMIENTO CON ESE NOMBRE YA EXISTE (" + fecha_nacimiento + ")" + '\n')

        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(exc_type, exc_obj, exc_tb.tb_lineno, flush=True)
            print('Formato Incorrecto')
            return {'message': 'Formato incorrecto', 'status': 400}, 400
        if len(errors) != 0:
            er = ''
            for e in errors:
                er += e + "\n"
            return {'message': er, 'status': 400}, 400
        else:
            return 'Estudiantes importados', 200